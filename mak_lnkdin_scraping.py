import os
import time
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException


url = Service()
driver = webdriver.Chrome(service=url)
driver.implicitly_wait(6)
driver.maximize_window()
driver.get("https://www.linkedin.com/")

# --- User Parameters ---
username = "username"
password = "password"
search_input = "drone"
location_search = "california"
location_select = "California, United States"

# --- Login ---
driver.find_element(By.XPATH, "//a[starts-with(@class, 'sign-in-form__sign-in-cta')]").click()
driver.find_element(By.ID,"username").send_keys(username)
driver.find_element(By.ID, "password").send_keys(password)
driver.find_element(By.XPATH, "//button[@type='submit']").click()

# --- Search and Location filter Set ---
wait = WebDriverWait(driver, 20)
wait.until(EC.visibility_of_element_located((By.XPATH, "//input[starts-with(@class, 'basic-input')]"))).send_keys(f"{search_input}" + Keys.ENTER)
wait.until(EC.visibility_of_element_located((By.XPATH, "//button[text() = 'Companies' and contains(@class, 'artdeco')]"))).click()
driver.find_element(By.XPATH, "//button[@id='searchFilter_companyHqGeo']//*[name()='svg']//*[name()='use' and contains(@href,'#caret-sma')]").click()
driver.find_element(By.XPATH, "//input[@placeholder='Add a location']").send_keys(location_search)
wait.until(EC.visibility_of_element_located((By.XPATH, f"//span[@class='search-typeahead-v2__hit-text t-14 t-black t-bold'][normalize-space() = '{location_select}']"))).click()
driver.find_element(By.XPATH, "//span[text() = 'Show results']").click()

data = []       # Empty list created

# --- scraper function defined ---
def scraper_func():
    for k in range(1, 12):
        try:
            time.sleep(2)
            iteration_list = driver.find_element(By.CSS_SELECTOR,
                                                 f"li:nth-child({k}) div:nth-child(1) div:nth-child(1) div:nth-child(1) div:nth-child(2) div:nth-child(1) div:nth-child(3)")
            iteration_list.click()
            time.sleep(2)
            driver.find_element(By.XPATH, "//a[text() = 'About']").click()

            def get_text_if_displayed(xpath):
                try:
                    element = driver.find_element(By.XPATH, xpath)
                    if element.is_displayed():
                        return element.text
                except NoSuchElementException:
                    pass
                return None

            company_name = get_text_if_displayed("//h1[contains(@class, 'summary__title')]")
            website = get_text_if_displayed("//dl[@class = 'overflow-hidden']/dd/a/span")
            industry = get_text_if_displayed("//h3[normalize-space()='Industry']//following::dd[1]")
            headquarter = get_text_if_displayed("// h3[normalize-space() = 'Headquarters'] // following::dd[1]")
            specialities = get_text_if_displayed("//dd[@dir = 'ltr']")

            data.append([company_name, website, industry, headquarter, specialities])

            driver.back()
            if driver.find_element(By.XPATH, "//a[text() = 'Home']").is_displayed():
                driver.back()
            elif driver.find_element(By.XPATH, "//span[text() = 'Follow']").is_displayed() is False:
                driver.back()

        except Exception as e:
            print(e)
            continue


scraper_func()      # Scraper function called for the first page

# --- Pagination ---
for i in range(2,4):
    try:
    #   wait.until(EC.visibility_of_element_located((By.XPATH, f"//li[@data-test-pagination-page-btn = '{i}']"))).click()
        wait.until(EC.visibility_of_element_located((By.XPATH, f"//span[@class = 'artdeco-button__text'][normalize-space() = 'Next']"))).click()
        scraper_func()
    except Exception as e:
        print(e)
        continue


# --------- Save to Excel file --------
if os.path.exists("linkedin_companies.xlsx"):
    wb = load_workbook("linkedin_companies.xlsx")  #  Load if file exists
    ws = wb.active
else:
    wb = Workbook()  # Only create new if it does not exist
    ws = wb.active
    ws.append(["Company Name", "Website", "Industry", "Headquarter", "Specialities"])  # Adding header in new file

for row in data:
    ws.append(row)

wb.save("linkedin_companies.xlsx")

